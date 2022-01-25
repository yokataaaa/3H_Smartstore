import requests
import time
import os
import openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import Workbook
from convert_xls_to_xlsx import convert_excel


def convert_all4lab_to_excel(url, new_file, category, mode):
    # 엑셀 파일 생성 or 열기
    # wb = Workbook(write_only=True)
    creat_file = new_file  # 생성 엑셀 파일명

    if mode == 'w':     # 새파일 생성
        convert_excel(creat_file)

    wb = openpyxl.load_workbook(creat_file)
    ws = wb.get_sheet_by_name('ver.2.1')

    # 엑셀 파일 마지막 행 찾기
    last_row = 0    # sheet의 마지막 행
    for cell in ws['A']:
        last_row += 1
        if cell.value == None:
            break

    # Headless 모드 온

    options = webdriver.ChromeOptions()
    options.add_argument('headless')
    options.add_argument('window-size=1920x1080')
    options.add_argument("disable-gpu")
    # options.add_argument("--disable-gpu")

    # 크롬 드라이버 생성
    driver = webdriver.Chrome(r'C:\Users\dlawp\Downloads\chromedriver_win32\chromedriver.exe', chrome_options=options)
    # 사이트 접속하기
    site_url = url  # 스크래핑 대상 url
    driver.get(site_url)
    driver.implicitly_wait(5)       # 페이지가 전부 로딩될까지 3초까지 기다린다

    html = driver.execute_script('return document.body.innerHTML')  # 셀레니움을 이용한 JavaScript => html 언어 크롤링


    # response = requests.get('https://www.allforlab.com/pdt/DH20030300P1776')

    # 한글 깨짐 파싱
    # soup = BeautifulSoup(response.content.decode('euc-kr', 'replace'), 'html.parser')
    soup = BeautifulSoup(html, 'html.parser')

    # 중요 옵션

    margin = 1.2   # 상품당 이익

    # 셀 서식

    A = '신상품'  # 상품상태 = ''   # A 필수(신상품/중고상품)
    B = 0  # 카테고리ID = '' # B 필수 DEFAULT = 50003439
    C = ''  # 상품명 = ''    # C 필수
    D = 99999  # 판매가 = ''    # D 필수(10원단위) int
    E = 0  # 재고수량 = ''   # E 필수 int
    F = '상품에 하자가 있거나 주문한 상품과 다를 경우, 수령일로부터 15일 이내인 경우에 한하여 1:1 상품교환 또는 전액 환불 조치 해드리며, ' \
        '이에 소요되는 모든 비용 또한 (주)쓰리에이치에서 부담합니다. ' \
        '' \
        '제품에 하자는 없지만, 다른 상품으로 교환하거나 반품하고자 할 경우, 출고일 기준 영업일 7일 이내, 유리 초자 제품은 3일이내에 반품하셔야 하며' \
        ' 사용하지 않은 새 제품인 경우에 한하여 조치해 드립니다. ' \
        '이에 소요되는 비용은 고객님 부담입니다. (제품 사용 후 변심에 의한 교환 및 환불은 불가)' \
        '' \
        '제품의 교환 및 반품 시에는 상품 비닐 및 박스(Brand Box)를 수령하신 상태 그대로 보존하고 있을 경우에만 가능합니다.' \
        '교환 및 반품 택배 보내실 때 상품 박스(Brand Box)가 손상되지 않도록 반드시 외부에 박스 한 겹 더 포장 부탁드립니다.' \
        '' \
        '건전지가 포함된 제품의 경우 수입 및 유통 과정에서 건전지의 전압 저하 및 방전의 경우가 있을 수 있습니다.' \
        '이럴 경우 영업 담당자에게 연락 주시면 신속하게 처리해 드리도록 하겠습니다.'  # A_S_안내내용 = '' # F 필수
    G = '010-5660-9934'  # A_S_전화번호 = '' # G 필수(02-0000-0000)
    H = ''  # 대표_이미지_파일명 = '' # H 필수
    I = ''  # 추가_이미지_파일명 = '' # I
    J = ''  # 상품_상세정보 = ''    # J 필수 ( 외부 이미지링크 )
    # 판매자_상품코드 = ''   # K
    # 판매자_바코드 = ''    # L
    M = ''  # 제조사 = ''    # M
    # 브랜드 = ''    # N
    # 제조일자 = ''   # O
    # 유효일자 = ''   # P
    Q = '과세상품'  # 부가세 = ''    # Q 필수(과세상품/면세상품/영세상품)
    R = 'Y'  # 미성년자_구매 = ''    # R 필수(Y/N)
    S = 'Y'  # 구매평_노출여부 = ''   # S 필수(Y/N)
    T = '00'  # 원산지_코드 = '' # T 필수
    # 수입사 = ''    # U
    # 복수원산지_여부 = ''   # V 필수(Y/N)
    # 원산지_직접입력 = ''   # W
    X = '택배, 소포, 등기'  # 배송방법 = ''   # X(택배,소포,등기/직접배송(화물배송) )
    Y = '유료'  # 배송비_유형 = '' # Y(무료/조건부 무료/유료/수량별)
    Z = 4000  # 기본배송비 = ''  # Z(배송비 유형이 무료 외에 필수 4000) int
    AA = '선결제'  # 배송비_결제방식 = ''   # AA(착불/선결제/착불 또는 선결제)
    # 조건부무료_상품판매가합계 = ''  # AB
    # 수량별부과_수량 = ''   # AC
    AD = 4000  # 반품배송비 = ''  # AD 조건부필수 4000   int
    AE = 4000  # 교환배송비 = ''  # AE 조건부필수 4000   int
    # 지역별_차등배송비_정보 = ''   # AF
    # 별도설치비 = ''  # AG
    # 판매자_특이사항 = ''   # AH
    # 즉시할인_값 = '' # AI
    # 즉시할인_단위 = ''    # AJ
    # 복수구매할인_조건_값 = ''    # AK
    # 복수구매할인_조건_단위 = ''   # AL
    # 복수구매할인_값 = ''   # AM
    # 복수구매할인_단위 = ''  # AN
    # 상품구매시_포인트_지급_값 = '' #AO
    # 상품구매시_포인트_지급_단위 = ''    # AP
    # 텍스트리뷰_작성시_지급_포인트 = ''   # AQ
    # 포토_동영상_리뷰_작성시_지급_포인트 = ''   # AR
    # 한달사용_텍스트리뷰_작성시_지급_포인트 = ''  # AS
    # 한달사용_포토_동영상리뷰_작성시_지급_포인트 = ''   # AT
    # 톡톡친구_스토어찜고객_리뷰_작성시_지급_포인트 = ''  # AU
    # 무이자_할부_개월 = ''  # AV
    # 사은품 = ''    # AW
    # 옵션형태 = ''   # AX
    # 옵션명 = ''    # AY
    # 옵션값 = ''    # AZ
    # 옵션가 = ''    # BA
    # 옵션_재고수량 = ''    # BB
    # 추가상품명 = ''  # BC
    # 추가상품값 = ''  # BD
    # 추가상품가 = ''  # BE
    # 추가상품_재고수량= ''   # BF
    # 상품정보제공고시_품명 = ''    # BG
    BH = ''     # 상품정보제공고시_모델명 = ''   # BH
    # 상품정보제공고시_인증허가사항 = ''    # BI
    # 상품정보제공고시_제조자 = ''   # BJ
    BK = 'N'  # 스토어찜회원_전용여부 = ''    # BK 필수(Y/N)
    # 문화비_소득공제 = ''   # BL
    # ISBN = ''   # BM
    # 독립출판 = ''   # BN
    # ISSN = ''   # BO


    # 이미지 파일 스크래핑

    maker = soup.select_one('#detailOpt > dd:nth-child(4)').get_text()  # 제조사
    # img_src = (soup.select_one('#mainimg')['onclick']).split("'")[1]
    img_paths = soup.select('#imgclick > li')


    try:
        detail_image = 'http:' + soup.select_one('#jump_0 > img')['src']  # 1장 상품 상세정보 이미지 파일 url
    except:
        if '<p>' in str(soup.select_one('#jump_0')):
            detail_image = 'http:' + soup.select_one('#jump_0 > p:nth-child(1) > img')['src']   # 2장이상 상품 상세정보 이미지 파일 url
        else:
            detail_image = ''

    img_srcs = list()  # 이미지 파일 경로
    temp_file_name = (site_url.split('/')[-1]).split('?')[0]
    img_file = list()  # 이미지 파일 이름 img_file[0] = 대표 이미지 img_file[1:] = 추가 이미지
    i = 0

    for img_path in img_paths:
        temp_img = img_path.select_one('img')['src']
        img_srcs.append(temp_img.replace('58', '500'))
        img_file.append(temp_file_name + '_{}'.format(i))
        i += 1

    # 이미지 파일 저장

    dir = 'image'
    i = 0
    filename = list()

    for i in range(len(img_srcs)):
        img_src = img_srcs[i]
        fname, extention = os.path.splitext(img_src)
        response = requests.get('https:' + img_src)
        filename.append(img_file[i] + extention)
        with open(os.path.join(dir, filename[i]), 'wb+') as f:
            f.write(response.content)

    # 테이블 정보 스크래핑

    tags = soup.select('.option_body > tr')#itemlist > div:nth-child(1) > table > tbody > tr:nth-child(1)
    i = last_row

    for tag in tags:
        td_tag = tag.select('td')

        cat_no = td_tag[0].get_text().strip()  # 카탈로그 넘버
        name = td_tag[1].get_text().strip()  # 상품명
        price = (td_tag[2].get_text().strip()).split(',')  # 소비자 가격
        unit = td_tag[5].get_text().strip()  # 단위
        inventory = ((td_tag[6].get_text()).split('(')[0]).strip()  # 재고수량
        B = category
        C = name
        D = round((int(''.join(price)))*margin, -2)

        if inventory == '단종':
            continue
        elif inventory == '∞':
            E = 999
        elif int(inventory) == 0:
            E = 1
        else:
            E = int(inventory)

        H = filename[0]
        I = ','.join(filename[1:])
        J = '<img src="' + detail_image + '">'
        M = maker
        BH = cat_no

        rows = [
            A,  # 상품상태 = ''   # A 필수(신상품/중고상품)
            B,  # 카테고리ID = '' # B 필수 int  50003439 생활-건강-기타측정기
            C,  # 상품명 = ''    # C 필수
            D,  # 판매가 = ''    # D 필수(10원단위) int
            E,  # 재고수량 = ''   # E 필수 int
            F,  # A_S_안내내용 = '' # F 필수
            G,  # A_S_전화번호 = '' # G 필수(02-0000-0000)
            H,  # 대표_이미지_파일명 = '' # H 필수
            I,  # 추가_이미지_파일명 = '' # I
            J,  # 상품_상세정보 = ''    # J 필수 ( 외부 이미지링크 )
            M,  # 제조사 = ''    # M
            Q,  # 부가세 = ''    # Q 필수(과세상품/면세상품/영세상품)
            R,  # 미성년자_구매 = ''    # R 필수(Y/N)
            S,  # 구매평_노출여부 = ''   # S 필수(Y/N)
            T,  # 원산지_코드 = '' # T 필수
            X,  # 배송방법 = ''   # X(택배,소포,등기/직접배송(화물배송) )
            Y,  # 배송비_유형 = '' # Y(무료/조건부 무료/유료/수량별)
            Z,  # 기본배송비 = ''  # Z(배송비 유형이 무료 외에 필수 2500) int
            AA,  # 배송비_결제방식 = ''   # AA(착불/선결제/착불 또는 선결제)
            AD,  # 반품배송비 = ''  # AD 조건부필수 2500   int
            AE,  # 교환배송비 = ''  # AE 조건부필수 5000   int
            BH,
            BK,  # 스토어찜회원_전용여부 = ''    # BK 필수(Y/N)
        ]

        # 엑셀에 저장
        # ws['A4'] = 1
        # ws.cell(2,2).value = 10
        ws['A' + str(i)] = A
        ws['B' + str(i)] = B
        ws['C' + str(i)] = C
        ws['D' + str(i)] = D
        ws['E' + str(i)] = E
        ws['F' + str(i)] = F
        ws['G' + str(i)] = G
        ws['H' + str(i)] = H
        ws['I' + str(i)] = I
        ws['J' + str(i)] = J
        ws['M' + str(i)] = M
        ws['Q' + str(i)] = Q
        ws['R' + str(i)] = R
        ws['S' + str(i)] = S
        ws['T' + str(i)] = T
        ws['X' + str(i)] = X
        ws['Y' + str(i)] = Y
        ws['Z' + str(i)] = Z
        ws['AA' + str(i)] = AA
        ws['AD' + str(i)] = AD
        ws['AE' + str(i)] = AE
        ws['BH' + str(i)] = BH
        ws['BK' + str(i)] = BK

        print(rows)
        i = i + 1

        # ws.append(rows)

    #
    driver.quit()
    wb.save(creat_file)
