import requests
import time
import os
import openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import Workbook

options = webdriver.ChromeOptions()
options.add_argument('headless')
options.add_argument('window-size=1920x1080')
options.add_argument("disable-gpu")
options.add_argument("--disable-gpu")
options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36")
options.add_argument("lang=ko_KR")

# 크롬 드라이버 생성
driver = webdriver.Chrome(r'C:\coding\chromedriver_win32\chromedriver.exe', chrome_options=options)
# 사이트 접속하기
site_url = 'http://www.lklab.com/product/product_info.asp?g_no=5444&t_no=780'  # 스크래핑 대상 url
driver.implicitly_wait(5)       # 페이지가 전부 로딩될까지 5초까지 기다린다
driver.get(site_url)
# html = driver.execute_script('return document.body.innerHTML')  # 셀레니움을 이용한 JavaScript => html inner body
html = driver.page_source   # page all
# response = requests.get('https://www.allforlab.com/pdt/DH20030300P1776')
# 한글 깨짐 파싱
# soup = BeautifulSoup(response.content.decode('euc-kr', 'replace'), 'html.parser')
soup = BeautifulSoup(html, 'html.parser')

index_list = list()
category_list = dict()
category_url = dict()
product_list = dict()
product_url = dict()

# list append ( A = id,   B = index,   C = category,   D = product,    E = url )
# append list [ A, B, C, D, E ]
# total list [ [ ], [ ], [ ] ]
row_list = list()

indexes = soup.select('#category_list ul')
indexes = indexes[0:1]
count = 0

for index in indexes:
    # print(index['id'])
    temp_index = index['id']
    index_list.append(temp_index)
    categorys = index.select('li a')
    temp_category_list = list()
    sub_url_list = list()

    for category in categorys:
        temp_category = category.text
        temp_category_list.append(temp_category)
        category_list.update({temp_index: temp_category_list})

        sub_url = 'http://www.lklab.com'+category['href']
        sub_url_list.append(sub_url)
        category_url.update({temp_index: sub_url_list})

        driver.get(sub_url)
        sub_html = driver.page_source
        sub_soup = BeautifulSoup(sub_html, 'html.parser')

        time.sleep(3)

        products = sub_soup.select('#content div.prod_box')
        temp_product_list = list()
        dst_url_list = list()

        for product in products:
            temp_product = product.select_one('p.name a').text
            temp_product_list.append(temp_product)
            product_list = {temp_category: temp_product_list}

            temp_url = product.select_one('p.name a')['href']
            temp_url = ''.join(temp_url[1:])
            dst_url = 'http://www.lklab.com/product' + temp_url
            dst_url_list.append(dst_url)
            product_url = {temp_category: dst_url_list}

            count += 1

            row = [
                count,
                temp_index,
                temp_category,
                temp_product,
                dst_url
            ]

            row_list.append(row)
            print(row)

print(count)


driver.quit()

# 엑셀 파일 생성 or 열기
wb = Workbook(write_only=True)
ws = wb.create_sheet('lklab')

ws.append(['id', 'index', 'Category', 'Product Name', 'URL'])

# A = id,   B = index,   C = category,   D = product,    E = url
for row in row_list:
    ws.append(row)

# row = 2
# for index in index_list:
#     ws['A'+str(row)] = row - 1
#     ws['B'+str(row)] = index
#     for category in category_list:
#         ws['C'+str(row)] = category
#         for i, product in enumerate(product_list):
#             ws['D'+str(row)] = product
#             ws['E'+str(row)] = product_url[i]
#             row += 1


wb.save('url_to_excel.xlsx')

