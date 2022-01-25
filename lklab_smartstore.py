import requests
import time
import os
import openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import Workbook
from pdf2image import convert_from_path
from convert_xls_to_xlsx import convert_excel

def convert_lklab_to_excel(url, new_file, category, mode):

    # 엑셀 파일 생성 or 열기
    # wb = Workbook(write_only=True)
    creat_file = new_file  # 생성 엑셀 파일명

    if mode == 's':     # 파일 초기화
        convert_excel(creat_file)
        return print('{} 파일 초기화 완료'.format(creat_file))
    elif mode == 'w':   # 덮어쓰기 모드( 새 파일 생성 후 )
        convert_excel(creat_file)
    else:
        pass

    # 이어쓰기 모드

    wb = openpyxl.load_workbook(creat_file)
    ws = wb.get_sheet_by_name('ver.2.1')

    # 엑셀 파일 마지막 행 찾기
    current_row = 0    # sheet의 마지막 행
    for cell in ws['A']:
        current_row += 1
        if cell.value == None:
            break
    print(current_row)

    if current_row >= 101:
        return False

    # Headless 모드 온

    options = webdriver.ChromeOptions()
    options.add_argument('headless')
    options.add_argument('window-size=1920x1080')
    options.add_argument("disable-gpu")
    options.add_argument("--disable-gpu")

    # Headless 모드 탐지 숨기기
    options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36")
    options.add_argument("lang=ko_KR") # 한국어!

    # 크롬 드라이버 생성
    driver = webdriver.Chrome(r'C:\coding\chromedriver_win32\chromedriver.exe', chrome_options=options)
    # 사이트 접속하기
    site_url = url  # 스크래핑 대상 url

    driver.implicitly_wait(5)       # 페이지가 전부 로딩될까지 5초까지 기다린다
    driver.get(site_url)

    # html = driver.execute_script('return document.body.innerHTML')  # 셀레니움을 이용한 JavaScript => html inner body
    html = driver.page_source   # page all


    # response = requests.get('https://www.allforlab.com/pdt/DH20030300P1776')

    # 한글 깨짐 파싱
    # soup = BeautifulSoup(response.content.decode('euc-kr', 'replace'), 'html.parser')
    soup = BeautifulSoup(html, 'html.parser')

    # 중요 옵션

    margin = 1   # 상품당 이익

    # 셀 서식

    A = '신상품'  # 상품상태 = ''   # A 필수(신상품/중고상품)
    B = category    # 카테고리ID = '' # B 필수 DEFAULT = 50003439
    C = ''  # 상품명 = ''    # C 필수
    D = 100  # 판매가 = ''    # D 필수(10원단위) int
    E = 99  # 재고수량 = ''   # E 필수 int
    F = '상품에 하자가 있거나 주문한 상품과 다를 경우, 수령일로부터 15일 이내인 경우에 한하여 1:1 상품교환 또는 전액 환불 조치 해드리며, ' \
        '이에 소요되는 모든 비용 또한 (주)쓰리에이치에서 부담합니다. ' \
        '' \
        '제품에 하자는 없지만, 다른 상품으로 교환하거나 반품하고자 할 경우, 출고일 기준 영업일 7일 이내, 유리 초자 제품은 3일이내에 반품하셔야 하며' \
        ' 사용하지 않은 새 제품인 경우에 한하여 조치해 드립니다. ' \
        '이에 소요되는 비용은 고객님 부담입니다. (제품 사용 후 변심에 의한 교환 및 환불은 불가)' \
        '' \
        '제품의 교환 및 반품 시에는 상품 비닐 및 박스(Brand Box)를 수령하신 상태 그대로 보존하고 있을 경우에만 가능합니다.' \
        '교환 및 반품 택배 보내실 때 상품 박스(Brand Box)가 손상되지 않도록 반드시 외부에 박스 한 겹 더 포장 부탁드립니다.' \
        '' \
        '건전지가 포함된 제품의 경우 수입 및 유통 과정에서 건전지의 전압 저하 및 방전의 경우가 있을 수 있습니다.' \
        '이럴 경우 영업 담당자에게 연락 주시면 신속하게 처리해 드리도록 하겠습니다.'  # A_S_안내내용 = '' # F 필수
    G = '010-5660-9934'  # A_S_전화번호 = '' # G 필수(02-0000-0000)
    H = ''  # 대표_이미지_파일명 = '' # H 필수
    I = ''  # 추가_이미지_파일명 = '' # I
    J = ''  # 상품_상세정보 = ''    # J 필수 ( 외부 이미지링크 )
    # 판매자_상품코드 = ''   # K
    # 판매자_바코드 = ''    # L
    # 제조사 = ''    # M
    # 브랜드 = ''    # N
    # 제조일자 = ''   # O
    # 유효일자 = ''   # P
    Q = '과세상품'  # 부가세 = ''    # Q 필수(과세상품/면세상품/영세상품)
    R = 'Y'  # 미성년자_구매 = ''    # R 필수(Y/N)
    S = 'Y'  # 구매평_노출여부 = ''   # S 필수(Y/N)
    T = '00'  # 원산지_코드 = '' # T 필수
    # 수입사 = ''    # U
    # 복수원산지_여부 = ''   # V 필수(Y/N)
    # 원산지_직접입력 = ''   # W
    X = '택배, 소포, 등기'  # 배송방법 = ''   # X(택배,소포,등기/직접배송(화물배송) )
    Y = '유료'  # 배송비_유형 = '' # Y(무료/조건부 무료/유료/수량별)
    Z = 4000  # 기본배송비 = ''  # Z(배송비 유형이 무료 외에 필수 4000) int
    AA = '선결제'  # 배송비_결제방식 = ''   # AA(착불/선결제/착불 또는 선결제)
    # 조건부무료_상품판매가합계 = ''  # AB
    # 수량별부과_수량 = ''   # AC
    AD = 4000  # 반품배송비 = ''  # AD 조건부필수 4000   int
    AE = 4000  # 교환배송비 = ''  # AE 조건부필수 4000   int
    # 지역별_차등배송비_정보 = ''   # AF
    # 별도설치비 = ''  # AG
    # 판매자_특이사항 = ''   # AH
    # 즉시할인_값 = '' # AI
    # 즉시할인_단위 = ''    # AJ
    # 복수구매할인_조건_값 = ''    # AK
    # 복수구매할인_조건_단위 = ''   # AL
    # 복수구매할인_값 = ''   # AM
    # 복수구매할인_단위 = ''  # AN
    # 상품구매시_포인트_지급_값 = '' #AO
    # 상품구매시_포인트_지급_단위 = ''    # AP
    # 텍스트리뷰_작성시_지급_포인트 = ''   # AQ
    # 포토_동영상_리뷰_작성시_지급_포인트 = ''   # AR
    # 한달사용_텍스트리뷰_작성시_지급_포인트 = ''  # AS
    # 한달사용_포토_동영상리뷰_작성시_지급_포인트 = ''   # AT
    # 톡톡친구_스토어찜고객_리뷰_작성시_지급_포인트 = ''  # AU
    # 무이자_할부_개월 = ''  # AV
    # 사은품 = ''    # AW
    # 옵션형태 = ''   # AX
    # 옵션명 = ''    # AY
    # 옵션값 = ''    # AZ
    # 옵션가 = ''    # BA
    # 옵션_재고수량 = ''    # BB
    BC = ''     # 추가상품명 = ''  # BC
    BD = ''     # 추가상품값 = ''  # BD
    BE = ''     # 추가상품가 = ''  # BE
    BF = ''     # 추가상품_재고수량= ''   # BF
    BG = ''     # 상품정보제공고시_품명 = ''    # BG
    BH = ''     # 상품정보제공고시_모델명 = ''   # BH
    # 상품정보제공고시_인증허가사항 = ''    # BI
    # 상품정보제공고시_제조자 = ''   # BJ
    BK = 'N'  # 스토어찜회원_전용여부 = ''    # BK 필수(Y/N)
    # 문화비_소득공제 = ''   # BL
    # ISBN = ''   # BM
    # 독립출판 = ''   # BN
    # ISSN = ''   # BO


    # 이미지 파일 스크래핑

    img_paths = soup.select('#thumb_s > ul > li')

    img_srcs = list()  # 이미지 파일 경로

    for img_path in img_paths:
        temp_img = 'http://www.lklab.com'+ img_path.select_one('img')['src']
        img_srcs.append(temp_img)

    # 이미지 파일 저장

    dir = 'image'
    detail_dir = 'detail_image'
    pdf_dir = 'pdf'

    i = 0
    filename = list()

    for i in range(len(img_srcs)):
        img_src = img_srcs[i]
        fname, extention = os.path.splitext(img_src)
        response = requests.get(img_src)
        temp_fname = fname.split('/')
        filename.append(temp_fname[-1]+extention)
        with open(os.path.join(dir, filename[i]), 'wb+') as f:
            f.write(response.content)

    # pdf_to_jpg 상세페이지
    try:
        pdf_path = soup.select_one('#prod_info_02 a')['href']
        pdf_src = 'http://www.lklab.com'+pdf_path
        pdf_file_name = pdf_dir+'/'+filename[0].split('.')[0]+'.pdf'
        response = requests.get(pdf_src)

        with open(pdf_file_name, 'wb+') as f:
            f.write(response.content)

        pages = convert_from_path(pdf_file_name)
        for i, page in enumerate(pages):
            pdf_to_jpg_fname = detail_dir+'/'+filename[0].split('.')[0]+'.jpg'
            page.save(pdf_to_jpg_fname, "JPEG")
    except:
        pass

    # html source 스크래핑

    src_link = 'src="http://www.lklab.com/'
    href_link = 'href="http://www.lklab.com'

    # head html

    head_html = '<meta http-equiv="Content-Type" content="text/html; charset=utf-8" />' \
                # '<link type="text/css" rel="stylesheet" href="http://www.lklab.com/../../css/common.css" />' \
                # '<link type="text/css" rel="stylesheet" href="http://www.lklab.com/../../css/content.css" />'

    # page 스타일시트 스크래핑 style_html

    style_html = ''
    tags = soup.select('style')
    for tag in tags:
        style_html = style_html + str(tag)      # tag list 를 한개의 str 로 변환

    style_html = style_html.replace('<style>', '')
    style_html = style_html.replace('</style>', '')
    style_html = '<style>'+style_html+'</style>'


    # 상품 프로필 스크래핑 profile_html

    profile_html = ''
    tags = soup.select('#prod_info > div')
    slice_tags = tags[:-1]    # 필요없는 부분 슬라이싱
    for tag in slice_tags:
        profile_html = profile_html + str(tag)
    profile_html = profile_html.replace('src="', src_link)
    profile_html = profile_html.replace('href="', href_link)

    eng_name = soup.select_one('#prod_info_01 > ul > li.name_eng').get_text().strip()   # 상품 영어명
    kor_name = soup.select_one('#prod_info_01 > ul > li.name_kor').get_text().strip()   # 상품 한국명

    # 상품 관련 자료 스크래핑 info_html

    info_html = ''
    if soup.select('#product_tab_01 > center') != None:
        tags = soup.select('#product_tab_01 > center')
        for tag in tags:
            info_html = info_html + str(tag)
        info_html = info_html.replace('src="', src_link)
        info_html = info_html.replace('href="', href_link+'/product/')

    # 상품 상세 규격 스크래핑 detail_html

    detail_html = ''
    if soup.select('#product_tab_02 > center') != None:
        tags = soup.select('#product_tab_02 > center')
        for tag in tags:
            detail_html = detail_html + str(tag)

    product_info_html = head_html+profile_html+info_html+detail_html     # head_html+style_html+



    # Css convert to inline CSS Style

    # /* standard table style */
    # .st-tablebox{ width:800px; max-width:800px; margin-bottom:10px; height:auto; border-top:#cdcdcd solid 3px;}
    # .st-tablebox p{ line-height: 19px;}
    # .st-tablebox th{ height:33px; background-color:#fbf8f4; font-size: 12px; color: #666; font-family:Tahoma; font-weight:bold; text-align:center;}
    # .st-tablebox td{ height:33px; font-size: 12px; color: #777; font-family:'돋움'; text-align:center;}
    # .st-tablebox th,td{border-left:#cdcdcd solid 1px; border-bottom:#cdcdcd solid 1px;}
    # .st-tablebox .tdfin{border-left:#cdcdcd solid 1px; border-right:#cdcdcd solid 1px;}
    # .st-tablebox .trfinth{background-color:#f5f7f1; font-size: 12px; color: #333; line-height: 33px; font-family:Tahoma; font-weight:bold; border-bottom:#cdcdcd solid 3px;  }
    # .st-tablebox .trfinth2{height:33px; background-color:#fbf8f4; font-size: 12px; color: #666; line-height: 33px; font-family:Tahoma; font-weight:bold; text-align:center; border-bottom:#cdcdcd solid 3px;  }
    # .st-tablebox .trfinth3{background-color:#f5f7f1; font-size: 12px; color: #333; line-height: 33px; font-family:Tahoma; font-weight:bold; border-bottom:#cdcdcd solid 1px;   border-right:#cdcdcd solid 1px;}
    # .st-tablebox .trfinth4{background-color:#f5f7f1; font-size: 12px; color: #333; line-height: 33px; font-family:Tahoma; font-weight:bold; border-bottom:#cdcdcd solid 1px;  }
    # .st-tablebox .trfin{border-bottom:#cdcdcd solid 3px; }
    # .st-tablebox .trfin2{border-bottom:#cdcdcd solid 3px; border-right:#cdcdcd solid 1px;}
    # .st-table_file td{border-left:#cdcdcd solid 0px; border-bottom:#cdcdcd solid 0px; }

    tableTrTd = 'margin: 0; padding: 0; border-collapse: collapse; font-size: 9pt; line-height: 1.3em; color: #666;'
    __sttablebox = 'width:800px; max-width:800px; margin-bottom:10px; height:auto; border-top:#cdcdcd solid 3px;'
    __sttablebox_p = 'line-height: 19px;'
    __sttablebox_th = 'height:33px; background-color:#fbf8f4; font-size: 12px; color: #666; font-family:Tahoma; font-weight:bold; text-align:center;'
    __sttablebox_td = 'height:33px; font-size: 12px; color: #777; font-family:'+'돋움'+'; text-align:center;'
    __sttablebox_thTd = 'border-left:#cdcdcd solid 1px; border-bottom:#cdcdcd solid 1px;'
    __sttablebox__tdfin = 'border-left:#cdcdcd solid 1px; border-right:#cdcdcd solid 1px;'
    __sttablebox__trfinth = 'background-color:#f5f7f1; font-size: 12px; color: #333; line-height: 33px; font-family:Tahoma; font-weight:bold; border-bottom:#cdcdcd solid 3px;'
    __sttablebox__trfinth2 = 'height:33px; background-color:#fbf8f4; font-size: 12px; color: #666; line-height: 33px; font-family:Tahoma; font-weight:bold; text-align:center; border-bottom:#cdcdcd solid 3px;'
    __sttablebox__trfinth3 = 'background-color:#f5f7f1; font-size: 12px; color: #333; line-height: 33px; font-family:Tahoma; font-weight:bold; border-bottom:#cdcdcd solid 1px;   border-right:#cdcdcd solid 1px;'
    __sttablebox__trfinth4 = 'background-color:#f5f7f1; font-size: 12px; color: #333; line-height: 33px; font-family:Tahoma; font-weight:bold; border-bottom:#cdcdcd solid 1px;'
    __sttablebox__trfin = 'border-bottom:#cdcdcd solid 3px;'
    __sttablebox__trfin2 = 'border-bottom:#cdcdcd solid 3px; border-right:#cdcdcd solid 1px;'
    __sttablefile_td = 'border-left:#cdcdcd solid 0px; border-bottom:#cdcdcd solid 0px;'

    style_context = 'style="'
    parser_html = product_info_html.split('>')
    context = list()

    for token in parser_html:
        if '<th ' in token:
            style_context += tableTrTd + __sttablebox_thTd + __sttablebox_th
            if 'class="tdfin"' in token:
                style_context += __sttablebox__tdfin + '"'
                token = token.replace('class="tdfin"', style_context)
            elif 'class="trfinth"' in token:
                style_context += __sttablebox__trfinth + '"'
                token = token.replace('class="trfinth"', style_context)
            elif 'class="trfinth2"' in token:
                style_context += __sttablebox__trfinth2 + '"'
                token = token.replace('class="trfinth2"', style_context)
            elif 'class="trfinth3"' in token:
                style_context += __sttablebox__trfinth3 + '"'
                token = token.replace('class="trfinth3"', style_context)
            elif 'class="trfinth4"' in token:
                style_context += __sttablebox__trfinth4 + '"'
                token = token.replace('class="trfinth4"', style_context)
            elif 'class="trfin"' in token:
                style_context += __sttablebox__trfin + '"'
                token = token.replace('class="trfin"', style_context)
            elif 'class="trfin2"' in token:
                style_context += __sttablebox__trfin2 + '"'
                token = token.replace('class="trfin2"', style_context)
            else:
                style_context += '"'
                token = token.replace('<th ', '<th ' + style_context)

        elif '<td ' in token:
            style_context += tableTrTd + __sttablebox_thTd + __sttablebox_td
            if 'class="tdfin"' in token:
                style_context += __sttablebox__tdfin + '"'
                token = token.replace('class="tdfin"', style_context)
            elif 'class="trfinth"' in token:
                style_context += __sttablebox__trfinth + '"'
                token = token.replace('class="trfinth"', style_context)
            elif 'class="trfinth2"' in token:
                style_context += __sttablebox__trfinth2 + '"'
                token = token.replace('class="trfinth2"', style_context)
            elif 'class="trfinth3"' in token:
                style_context += __sttablebox__trfinth3 + '"'
                token = token.replace('class="trfinth3"', style_context)
            elif 'class="trfinth4"' in token:
                style_context += __sttablebox__trfinth4 + '"'
                token = token.replace('class="trfinth4"', style_context)
            elif 'class="trfin"' in token:
                style_context += __sttablebox__trfin + '"'
                token = token.replace('class="trfin"', style_context)
            elif 'class="trfin2"' in token:
                style_context += __sttablebox__trfin2 + '"'
                token = token.replace('class="trfin2"', style_context)
            else:
                style_context += '"'
                token = token.replace('<td ', '<td ' + style_context)

        context.append(token)
        style_context = 'style="'

    context = '>'.join(context)

    # 테이블
    context = context.replace('class="st-tablebox"', style_context+tableTrTd+__sttablebox+'"')
    context = context.replace('td class="st-table_file"', 'td '+style_context+tableTrTd+__sttablefile_td+'"')
    # table > p
    context = context.replace('p scope="row"', 'p '+style_context+tableTrTd+__sttablebox_p+'"')
    # # table > th
    # product_info_html = product_info_html.replace('th scope="row"', 'th '+style_context+tableTrTd+__sttablebox_th+__sttablebox_thTd+'"')
    # product_info_html = product_info_html.replace('th class="tdfin"', 'th '+style_context+tableTrTd+__sttablebox_th+__sttablebox_thTd+__sttablebox__tdfin+'"')
    # product_info_html = product_info_html.replace('th class="trfinth"', 'th '+style_context+tableTrTd+__sttablebox_th+__sttablebox_thTd+__sttablebox__trfinth+'"')
    # product_info_html = product_info_html.replace('th class="trfinth2"', 'th '+style_context+tableTrTd+__sttablebox_th+__sttablebox_thTd+__sttablebox__trfinth2+'"')
    # product_info_html = product_info_html.replace('th class="trfinth3"', 'th '+style_context+tableTrTd+__sttablebox_th+__sttablebox_thTd+__sttablebox__trfinth3+'"')
    # product_info_html = product_info_html.replace('th class="trfinth4"', 'th '+style_context+tableTrTd+__sttablebox_th+__sttablebox_thTd+__sttablebox__trfinth4+'"')
    # product_info_html = product_info_html.replace('th class="trfin"', 'th '+style_context+tableTrTd+__sttablebox_th+__sttablebox_thTd+__sttablebox__trfin+'"')
    # product_info_html = product_info_html.replace('th class="trfin2"', 'th '+style_context+tableTrTd+__sttablebox_th+__sttablebox_thTd+__sttablebox__trfin2+'"')
    # # table > td
    # product_info_html = product_info_html.replace('td scope="row"', 'td '+style_context+tableTrTd+__sttablebox_td+__sttablebox_thTd+'"')
    # product_info_html = product_info_html.replace('td class="tdfin"', 'td '+style_context+tableTrTd+__sttablebox_td+__sttablebox_thTd+__sttablebox__tdfin+'"')
    # product_info_html = product_info_html.replace('td class="trfinth"', 'td '+style_context+tableTrTd+__sttablebox_td+__sttablebox_thTd+__sttablebox__trfinth+'"')
    # product_info_html = product_info_html.replace('td class="trfinth2"', 'td '+style_context+tableTrTd+__sttablebox_td+__sttablebox_thTd+__sttablebox__trfinth2+'"')
    # product_info_html = product_info_html.replace('td class="trfinth3"', 'td '+style_context+tableTrTd+__sttablebox_td+__sttablebox_thTd+__sttablebox__trfinth3+'"')
    # product_info_html = product_info_html.replace('td class="trfinth4"', 'td '+style_context+tableTrTd+__sttablebox_td+__sttablebox_thTd+__sttablebox__trfinth4+'"')
    # product_info_html = product_info_html.replace('td class="trfin"', 'td '+style_context+tableTrTd+__sttablebox_td+__sttablebox_thTd+__sttablebox__trfin+'"')
    # product_info_html = product_info_html.replace('td class="trfin2"', 'td '+style_context+tableTrTd+__sttablebox_td+__sttablebox_thTd+__sttablebox__trfin2+'"')
    # 제목
    context = context.replace('"http://www.lklab.com/../../images/product/img_substance.png"','"https://raw.githubusercontent.com/yokataaaa/3H_Smartstore/main/image/img_substance.png"')
    context = context.replace('id="prod_info_01"', 'style="width:550px; margin-top:30px;"')
    context = context.replace('class="keyword"', 'style="list-style:none; margin:0; padding:0; color:#666; font-size:9pt; line-height:1.3em; letter-spacing:-0.05em;color:#333; margin-top:15px; margin-bottom:25px; line-height:2.2em;"')
    context = context.replace('id="prod_info_02"', 'style="width:550px; margin-top:35px;"')
    context = context.replace('class="name_eng"', 'style="list-style:none; margin:0; padding:0; color:#666; line-height:1.3em; letter-spacing:-0.05em; font-size:15pt; margin-bottom:40px; width:550px; margin-top:30px;"')
    context = context.replace('class="name_kor"', 'style="list-style:none; margin:0; padding:0; color:#666; line-height:1.3em; letter-spacing:-0.05em; font-size:15pt; margin-bottom:40px; width:550px; margin-top:30px;"')

    # html test

    with open('test.html', 'w', encoding='UTF-8') as file:
        file.write(context)              # Test Html Code


    # 상품 주문 스크래핑
    # iframe html 스크래핑  driver.switch_to.frame("id 또는 name")

    iframe_src = soup.select_one('iframe')['id']    # iframe id 서칭 저장

    # iframe html 으로 전환
    driver.switch_to.frame(iframe_src)
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')

    tags = soup.select('#Table2 > tbody > tr')
    tr_tags = tags[1:]

    # 추가 상품 Description 내용 스크래핑
    cat_no = list()
    model = list()
    description = list()
    unit = list()
    price = list()
    inventory = list()
    count = 0       # 추가상품 갯수

    for tag in tr_tags:
        td_tag = tag.select('td')
        if len(td_tag) > 1:
            cat_no.append(td_tag[1].text.strip())     # Cat. No
            model.append(td_tag[3].text.strip())     # 모델명

            des_40 = td_tag[5].contents[0].strip()
            des_40 = str(des_40).replace('>', '-')
            if len(des_40) >= 40:
                description.append(''.join(des_40[:40]))
            else:
                description.append(des_40)      # 추가상품 Description


            unit.append(td_tag[7].text.split('/')[0])     # 단위

            try:
                cost = (''.join(td_tag[9].text.split(':')[1].strip().split(',')).replace('원', ''))    # 가격 ( VAT 미포함 )
                price.append(str(round((int(cost))*margin, -2)))
            except:
                price.append('999900')

            # 재고 ( 본사 contents[0] 공장 contents[2] ) * 예외처리
            if len(td_tag[11].contents) > 1:
                company_inventory = td_tag[11].contents[0].split(':')[1].strip()   # 본사 재고
                factory_inventory = td_tag[11].contents[2].split(':')[1].strip()   # 공장 재고
                inventory.append(str(int(company_inventory)+int(factory_inventory)))
            else:
                inventory.append('1')  # 예외처리 15 day, 30 day

            count += 1

    print(count)
    # # convert to inline CSS with Selenium
    #
    # convert_css_url = 'https://htmlemail.io/inline/'
    # driver.get(convert_css_url)
    # html = driver.page_source
    # soup = BeautifulSoup(html, 'html.parser')
    #
    # time.sleep(0.5)
    # driver.find_element_by_css_selector('#input').clear()
    # time.sleep(0.5)
    #
    # # long_string= <the long string>
    # # input_box = driver.find_element_by_id('translateText')
    # # driver.execute_script('arguments[0].value=arguments[1]', input_box, long_string) 대량의 string 입력




    # 스크래핑 자료 엑셀 저장

    C = eng_name+'('+kor_name+')'
    D = 100
    E = 99

    H = filename[0]
    I = ','.join(filename[1:])
    J = context+'<br>'+'<img src="https://www.pythonanywhere.com/user/principe84/files/home/principe84/detail_image/'+filename[0].split('.')[0]+'.jpg">'

    # ex) ws['A4'] = 1
    # ex) ws.cell(4,1).value = 1

    if count >= 10:
        index = current_row
        BC = '\n'.join(cat_no[:10])  # 추가상품명 = ''  # BC
        BD = ',\n'.join(description[:10])  # 추가상품값 = ''  # BD
        BE = ',\n'.join(price[:10])  # 추가상품가 = ''  # BE
        BF = ',\n'.join(inventory[:10])  # 추가상품_재고수량= ''   # BF
        BG = ''  # 상품정보제공고시_품명 = ''    # BG
        BH = ''  # 상품정보제공고시_모델명 = ''   # BH

        ws['A' + str(index)] = A
        ws['B' + str(index)] = B
        ws['C' + str(index)] = C
        ws['D' + str(index)] = D
        ws['E' + str(index)] = E
        ws['F' + str(index)] = F
        ws['G' + str(index)] = G
        ws['H' + str(index)] = H
        ws['I' + str(index)] = I
        ws['J' + str(index)] = J
        ws['Q' + str(index)] = Q
        ws['R' + str(index)] = R
        ws['S' + str(index)] = S
        ws['T' + str(index)] = T
        ws['X' + str(index)] = X
        ws['Y' + str(index)] = Y
        ws['Z' + str(index)] = Z
        ws['AA' + str(index)] = AA
        ws['AD' + str(index)] = AD
        ws['AE' + str(index)] = AE
        ws['BC' + str(index)] = BC
        ws['BD' + str(index)] = BD
        ws['BE' + str(index)] = BE
        ws['BF' + str(index)] = BF
        ws['BG' + str(index)] = BG
        ws['BH' + str(index)] = BH
        ws['BK' + str(index)] = BK

        index = current_row + 1
        BC = '\n'.join(cat_no[10:])  # 추가상품명 = ''  # BC
        BD = ',\n'.join(description[10:])  # 추가상품값 = ''  # BD
        BE = ',\n'.join(price[10:])  # 추가상품가 = ''  # BE
        BF = ',\n'.join(inventory[10:])  # 추가상품_재고수량= ''   # BF
        BG = ''  # 상품정보제공고시_품명 = ''    # BG
        BH = ''  # 상품정보제공고시_모델명 = ''   # BH

        ws['A' + str(index)] = A
        ws['B' + str(index)] = B
        ws['C' + str(index)] = C
        ws['D' + str(index)] = D
        ws['E' + str(index)] = E
        ws['F' + str(index)] = F
        ws['G' + str(index)] = G
        ws['H' + str(index)] = H
        ws['I' + str(index)] = I
        ws['J' + str(index)] = J
        ws['Q' + str(index)] = Q
        ws['R' + str(index)] = R
        ws['S' + str(index)] = S
        ws['T' + str(index)] = T
        ws['X' + str(index)] = X
        ws['Y' + str(index)] = Y
        ws['Z' + str(index)] = Z
        ws['AA' + str(index)] = AA
        ws['AD' + str(index)] = AD
        ws['AE' + str(index)] = AE
        ws['BC' + str(index)] = BC
        ws['BD' + str(index)] = BD
        ws['BE' + str(index)] = BE
        ws['BF' + str(index)] = BF
        ws['BG' + str(index)] = BG
        ws['BH' + str(index)] = BH
        ws['BK' + str(index)] = BK

    else:
        index = current_row
        BC = '\n'.join(cat_no)  # 추가상품명 = ''  # BC
        BD = ',\n'.join(description)  # 추가상품값 = ''  # BD
        BE = ',\n'.join(price)  # 추가상품가 = ''  # BE
        BF = ',\n'.join(inventory)  # 추가상품_재고수량= ''   # BF
        BG = ''  # 상품정보제공고시_품명 = ''    # BG
        BH = ''  # 상품정보제공고시_모델명 = ''   # BH

        ws['A' + str(index)] = A
        ws['B' + str(index)] = B
        ws['C' + str(index)] = C
        ws['D' + str(index)] = D
        ws['E' + str(index)] = E
        ws['F' + str(index)] = F
        ws['G' + str(index)] = G
        ws['H' + str(index)] = H
        ws['I' + str(index)] = I
        ws['J' + str(index)] = J
        ws['Q' + str(index)] = Q
        ws['R' + str(index)] = R
        ws['S' + str(index)] = S
        ws['T' + str(index)] = T
        ws['X' + str(index)] = X
        ws['Y' + str(index)] = Y
        ws['Z' + str(index)] = Z
        ws['AA' + str(index)] = AA
        ws['AD' + str(index)] = AD
        ws['AE' + str(index)] = AE
        ws['BC' + str(index)] = BC
        ws['BD' + str(index)] = BD
        ws['BE' + str(index)] = BE
        ws['BF' + str(index)] = BF
        ws['BG' + str(index)] = BG
        ws['BH' + str(index)] = BH
        ws['BK' + str(index)] = BK

    rows = [
        A,  # 상품상태 = ''   # A 필수(신상품/중고상품)
        B,  # 카테고리ID = '' # B 필수 int  50003439 생활-건강-기타측정기
        C,  # 상품명 = ''    # C 필수
        D,  # 판매가 = ''    # D 필수(10원단위) int
        E,  # 재고수량 = ''   # E 필수 int
        F,  # A_S_안내내용 = '' # F 필수
        G,  # A_S_전화번호 = '' # G 필수(02-0000-0000)
        H,  # 대표_이미지_파일명 = '' # H 필수
        I,  # 추가_이미지_파일명 = '' # I
        J,  # 상품_상세정보 = ''    # J 필수 ( 외부 이미지링크 )
        Q,  # 부가세 = ''    # Q 필수(과세상품/면세상품/영세상품)
        R,  # 미성년자_구매 = ''    # R 필수(Y/N)
        S,  # 구매평_노출여부 = ''   # S 필수(Y/N)
        T,  # 원산지_코드 = '' # T 필수
        X,  # 배송방법 = ''   # X(택배,소포,등기/직접배송(화물배송) )
        Y,  # 배송비_유형 = '' # Y(무료/조건부 무료/유료/수량별)
        Z,  # 기본배송비 = ''  # Z(배송비 유형이 무료 외에 필수 2500) int
        AA,  # 배송비_결제방식 = ''   # AA(착불/선결제/착불 또는 선결제)
        AD,  # 반품배송비 = ''  # AD 조건부필수 2500   int
        AE,  # 교환배송비 = ''  # AE 조건부필수 5000   int
        BC,  # 추가상품명 = ''  # BC
        BD,  # 추가상품값 = ''  # BD
        BE,  # 추가상품가 = ''  # BE
        BF,  # 추가상품_재고수량= ''   # BF
        BG,  # 상품정보제공고시_품명 = ''    # BG
        BH,  # 상품정보제공고시_모델명 = ''   # BH
        BK,  # 스토어찜회원_전용여부 = ''    # BK 필수(Y/N)
    ]

    # test print(rows)

    #
    driver.quit()
    wb.save(creat_file)

    return True

# test

# url = 'http://www.lklab.com/product/product_info.asp?g_no=13275&t_no=780'
# new_file = 'test.xlsx'
# category = 50003349
# mode = 'r'
#
# convert_lklab_to_excel(url, new_file, category, mode)
